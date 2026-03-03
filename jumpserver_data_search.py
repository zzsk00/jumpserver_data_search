import pymysql
import warnings
from datetime import datetime
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import sys
import os

# 忽略MySQL警告
warnings.filterwarnings('ignore')


class JumpServerDBQuery:
    def __init__(self):
        # 数据库配置
        self.config = {
            "host": "10.1.14.133",
            "port": 3306,
            "user": "root",
            "password": "Root@123456",
            "db": "jumpserver",
            "charset": "utf8mb4",
            "cursorclass": pymysql.cursors.DictCursor
        }
        self.connection = None
        self.results = {}

        # ===================== 你要的菜单 =====================
        self.query_menu = {
            "1":  {"name": "近30天活跃用户数",            "type": "sql",   "sql": "SELECT COUNT(DISTINCT username) AS active_user_count FROM audits_userloginlog WHERE TO_DAYS(NOW()) - TO_DAYS(datetime) <= 30;"},
            "2":  {"name": "近30天会话数",                "type": "sql",   "sql": "SELECT COUNT(*) AS session_count FROM terminal_session WHERE DATEDIFF(NOW(), date_start) <= 30;"},
            "3":  {"name": "近30天命令记录数",            "type": "sql",   "sql": "SELECT COUNT(*) AS command_count FROM terminal_command WHERE TO_DAYS(NOW()) - TO_DAYS(FROM_UNIXTIME(timestamp)) <= 30;"},
            "4":  {"name": "近30天资产登录次数",          "type": "sql",   "sql": "SELECT COUNT(DISTINCT asset) AS asset_count FROM terminal_session WHERE DATEDIFF(NOW(), date_end) <= 30;"},
            "5":  {"name": "近30天操作日志数量",          "type": "sql",   "sql": "SELECT COUNT(*) AS op_count FROM audits_operatelog WHERE DATEDIFF(NOW(), datetime) <= 30 AND user <> 'cmpSyncUser(cmpSyncUser)';"},
            "6":  {"name": "近30天登录日志数量",          "type": "sql",   "sql": "SELECT COUNT(*) AS login_log_count FROM audits_userloginlog WHERE DATEDIFF(NOW(), datetime) <= 30;"},
            "7":  {"name": "近30天登录数",                "type": "sql",   "sql": "SELECT COUNT(*) AS total_login_count FROM audits_userloginlog WHERE TO_DAYS(NOW()) - TO_DAYS(datetime) <= 30;"},
            "8":  {"name": "近1天登录次数",               "type": "sql",   "sql": "SELECT username, COUNT(*) AS login_count FROM audits_userloginlog WHERE TO_DAYS(NOW()) - TO_DAYS(datetime) <= 1 GROUP BY username ORDER BY login_count DESC;"},
            "9":  {"name": "近1天资产访问次数",           "type": "sql",   "sql": "SELECT asset, COUNT(*) AS cnt FROM terminal_session WHERE TO_DAYS(NOW()) - TO_DAYS(date_end) <=1 GROUP BY asset ORDER BY cnt DESC;"},
            "10": {"name": "近1月登录次数",              "type": "sql",   "sql": "SELECT username, COUNT(*) AS login_count FROM audits_userloginlog WHERE TO_DAYS(NOW()) - TO_DAYS(datetime) <=30 GROUP BY username ORDER BY login_count DESC;"},
            "11": {"name": "近一周用户登录次数",          "type": "sql",   "sql": "SELECT username, COUNT(*) AS login_count FROM audits_userloginlog WHERE TO_DAYS(NOW()) - TO_DAYS(datetime) <=7 GROUP BY username ORDER BY login_count DESC;"},
            "12": {"name": "近一周资产访问次数",          "type": "sql",   "sql": "SELECT asset, COUNT(*) AS cnt FROM terminal_session WHERE TO_DAYS(NOW()) - TO_DAYS(date_end) <=7 GROUP BY asset ORDER BY cnt DESC;"},
            "13": {"name": "近30天上传文件数量",          "type": "sql",   "sql": "SELECT COUNT(*) AS ftp_count FROM audits_ftplog WHERE DATEDIFF(NOW(), date_start) <=30;"},
            "14": {"name": "当前在线用户",                "type": "sql",   "sql": "SELECT * FROM terminal_session WHERE is_finished = '0';"},
            "15": {"name": "近30天连接过的资产",          "type": "sql",   "sql": "SELECT DISTINCT SUBSTRING_INDEX(SUBSTRING_INDEX(asset,'(',-1),')',1) AS ip, SUBSTRING_INDEX(asset,'(',1) AS name FROM terminal_session WHERE date_start >= DATE_SUB(NOW(),INTERVAL 30 DAY) ORDER BY ip;"},
            "16": {"name": "某用户的资产数量",            "type": "user",  "sql": "SELECT user_id, user_name, COUNT(asset_id) AS asset_count FROM (SELECT uu.id AS user_id, uu.name AS user_name, pau.assetpermission_id FROM users_user uu LEFT JOIN perms_assetpermission_users pau ON uu.id=pau.user_id WHERE uu.role!='App' AND uu.created_by!='System' UNION SELECT us.id user_id, us.name user_name, paug.assetpermission_id FROM perms_assetpermission_user_groups paug JOIN users_usergroup uu ON paug.usergroup_id=uu.id JOIN users_user_groups uug ON uu.id=uug.usergroup_id JOIN users_user us ON us.id=uug.user_id) main_1 LEFT JOIN (SELECT paa.assetpermission_id, paa.asset_id FROM perms_assetpermission_assets paa UNION SELECT pan.assetpermission_id, aan.asset_id FROM perms_assetpermission_nodes pan JOIN assets_asset_nodes aan ON pan.node_id=aan.node_id) main_2 ON main_1.assetpermission_id=main_2.assetpermission_id WHERE user_name=%s GROUP BY user_id, user_name;"},
            "17": {"name": "某组织资源数量",              "type": "org",   "sql": "SELECT t.NAME, COUNT(*) AS server_count FROM assets_asset a, orgs_organization t WHERE LEFT(a.org_id,8)=LEFT(t.id,8) AND t.NAME=%s GROUP BY t.NAME;"},
            # ========== 新增：18. 某一资产授权的用户名 ==========
            "18": {"name": "某一资产授权的用户名",         "type": "asset","sql": "SELECT DISTINCT aa.name AS 资产名称, uu.name AS 授权用户名 FROM perms_assetpermission_assets paa INNER JOIN perms_assetpermission_users pau ON paa.assetpermission_id = pau.assetpermission_id INNER JOIN users_user uu ON pau.user_id = uu.id INNER JOIN assets_asset aa ON paa.asset_id = aa.id WHERE aa.name = %s;"},        }

    def connect(self):
        try:
            self.connection = pymysql.connect(**self.config)
            print(f"✅ 连接成功：{self.config['host']}")
            return True
        except Exception as e:
            print(f"❌ 连接失败：{e}")
            return False

    def exe(self, sql, name, params=None):
        if not self.connection:
            print("❌ 未连接")
            return []
        try:
            with self.connection.cursor() as cur:
                cur.execute(sql, params or ())
                res = cur.fetchall()
                print(f"✅ {name} → {len(res)} 条")
                return res
        except Exception as e:
            print(f"❌ 查询失败：{e}")
            return []

    def show_menu(self):
        os.system('cls' if os.name == 'nt' else 'clear')
        print("=" * 90)
        print("📋 JumpServer 统计查询工具")
        print("=" * 90)
        for k in sorted(self.query_menu.keys(), key=lambda x: int(x)):
            print(f" {k:>2}. {self.query_menu[k]['name']}")
        print("=" * 90)
        print("操作：输入数字｜回车/tab 返回菜单｜q 退出")
        print("=" * 90)

    def input_(self, prompt):
        try:
            i = input(prompt).strip().lower()
            if i == "" or i == "tab":
                return "tab"
            return i
        except:
            return "tab"

    def export_excel(self, data, name):
        if not data:
            print("❌ 无数据")
            return
        name = name.replace("/","").replace("\\","").replace(":","")
        f = f"jumpserver_{name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = name[:30]

        # 表头样式
        hf = Font(bold=True, color="FFFFFF")
        hb = PatternFill("solid", "4472C4", "4472C4")
        ha = Alignment(horizontal="center", vertical="center")

        headers = list(data[0].keys())
        for c, h in enumerate(headers,1):
            cell = ws.cell(1,c,h)
            cell.font = hf
            cell.fill = hb
            cell.alignment = ha

        for r, row in enumerate(data,2):
            for c, k in enumerate(headers,1):
                v = row.get(k,"")
                if isinstance(v, datetime):
                    v = v.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(r,c,str(v))

        # 列宽
        for col in ws.columns:
            l = max((len(str(v.value)) if v.value else 0) for v in col)
            ws.column_dimensions[col[0].column_letter].width = min(l+2,60)

        wb.save(f)
        wb.close()
        print(f"💾 已导出：{f}")

    def run(self):
        while True:
            self.show_menu()
            c = self.input_("请输入编号：")
            if c == "q":
                print("👋 退出")
                break
            if c == "tab":
                continue
            if c not in self.query_menu:
                input("无效编号，按回车返回")
                continue

            item = self.query_menu[c]
            name = item["name"]
            sql = item["sql"]
            typ = item.get("type","sql")
            params = None

            print(f"\n🔍 {name}")

            # ========== 新增：处理asset类型的参数输入 ==========
            if typ == "asset":
                asset_name = self.input_("输入资产名称：")
                if asset_name == "tab": continue
                params = (asset_name,)
            elif typ == "user":
                u = self.input_("输入用户名：")
                if u == "tab": continue
                params = (u,)
            elif typ == "org":
                o = self.input_("输入组织名称：")
                if o == "tab": continue
                params = (o,)

            data = self.exe(sql, name, params)
            if not data:
                input("\n无数据，按回车返回")
                continue

            # 展示
            print("\n📊 结果预览：")
            for i, row in enumerate(data[:10],1):
                print(f"{i:2}. {row}")
            if len(data) > 10:
                print(f"... 共 {len(data)} 条")

            # 导出
            e = self.input_("\n导出 Excel ? (y/n)：")
            if e == "y":
                self.export_excel(data, name)

            input("\n按回车返回菜单")

    def close(self):
        if self.connection:
            self.connection.close()
            print("\n🔌 连接已关闭")


if __name__ == "__main__":
    app = JumpServerDBQuery()
    if not app.connect():
        input("按回车退出")
        sys.exit(1)
    try:
        app.run()
    except Exception as e:
        print(f"\n❌ 错误：{e}")
    app.close()